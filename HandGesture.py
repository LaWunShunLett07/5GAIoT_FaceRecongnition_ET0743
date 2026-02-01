#!/usr/bin/env python3
"""
🤚 Simple Hand Gesture Detection (OpenCV Only - No MediaPipe)
Uses skin detection and contour analysis

Detects 3 gestures:
1. ✋ OPEN HAND (5 fingers)
2. ✌️ PEACE SIGN (2 fingers)
3. 👊 FIST (closed hand)

Press 'q' to quit
"""

import cv2
import numpy as np
import time
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from imutils.video import VideoStream
import os

# ==============================================================================
# ⚙️ SETTINGS
# ==============================================================================
RTSP_URL = "rtsp://192.168.0.10:8554/feeder"  # ← Change to your Pi IP
USE_WEBCAM = False  # Set True to use laptop webcam

# Display
DISPLAY_WIDTH = 1000
DISPLAY_HEIGHT = 800

# Excel logging
EXCEL_FILE = "simple_gesture_log.xlsx"
LOG_EVERY_N_SECONDS = 3
# ==============================================================================

os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = "rtsp_transport;tcp|fflags;nobuffer|flags;low_delay"


# ==============================================================================
# 📊 EXCEL LOGGING
# ==============================================================================
def init_excel_log():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Gesture Log"
        
        headers = ["Timestamp", "Gesture Detected", "Finger Count", "Notes"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        
        wb.save(EXCEL_FILE)
        print(f"✅ Excel: {EXCEL_FILE}")

def log_to_excel(gesture, finger_count):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        notes = f"Detected {finger_count} fingers" if finger_count > 0 else "No hand detected"
        
        row = [timestamp, gesture, finger_count, notes]
        ws.append(row)
        
        wb.save(EXCEL_FILE)
        print(f"📊 Log: {gesture} ({finger_count} fingers)")
    except Exception as e:
        print(f"❌ Log failed: {e}")


# ==============================================================================
# 🤚 SIMPLE GESTURE RECOGNITION
# ==============================================================================
class SimpleGestureDetector:
    def __init__(self):
        # Skin color range in HSV (adjust for your skin tone)
        self.lower_skin = np.array([0, 20, 70], dtype=np.uint8)
        self.upper_skin = np.array([20, 255, 255], dtype=np.uint8)
    
    def detect_skin(self, frame):
        """Detect skin-colored regions"""
        # Convert to HSV
        hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        
        # Threshold for skin
        mask = cv2.inRange(hsv, self.lower_skin, self.upper_skin)
        
        # Clean up noise
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (11, 11))
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
        
        # Blur
        mask = cv2.GaussianBlur(mask, (3, 3), 0)
        
        return mask
    
    def count_fingers(self, contour, defects):
        """Count extended fingers using convexity defects"""
        if defects is None:
            return 0
        
        finger_count = 0
        
        for i in range(defects.shape[0]):
            s, e, f, d = defects[i, 0]
            start = tuple(contour[s][0])
            end = tuple(contour[e][0])
            far = tuple(contour[f][0])
            
            # Calculate lengths of triangle sides
            a = np.sqrt((end[0] - start[0])**2 + (end[1] - start[1])**2)
            b = np.sqrt((far[0] - start[0])**2 + (far[1] - start[1])**2)
            c = np.sqrt((end[0] - far[0])**2 + (end[1] - far[1])**2)
            
            # Calculate angle
            angle = np.arccos((b**2 + c**2 - a**2) / (2 * b * c))
            
            # If angle < 90 degrees, count as a finger gap
            if angle <= np.pi / 2:
                finger_count += 1
        
        # Add 1 because finger count = defects + 1
        return finger_count + 1 if finger_count > 0 else 0
    
    def recognize_gesture(self, finger_count):
        """Convert finger count to gesture name"""
        if finger_count == 0:
            return "👊 FIST"
        elif finger_count == 1:
            return "☝️ ONE FINGER"
        elif finger_count == 2:
            return "✌️ PEACE"
        elif finger_count >= 5:
            return "✋ OPEN HAND"
        else:
            return f"🤚 {finger_count} FINGERS"


# ==============================================================================
# 🎨 UI DRAWING
# ==============================================================================
def draw_info(frame, gesture, finger_count, fps):
    h, w = frame.shape[:2]
    
    # Semi-transparent overlay
    overlay = frame.copy()
    cv2.rectangle(overlay, (10, 10), (w - 10, 150), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.4, frame, 0.6, 0, frame)
    
    # Timestamp
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    cv2.putText(frame, f"🕒 {timestamp}", (20, 40), 
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
    
    # Gesture info
    cv2.putText(frame, f"Gesture: {gesture}", (20, 75),
                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)
    
    cv2.putText(frame, f"Fingers: {finger_count}", (20, 110),
                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 200, 0), 2)
    
    # FPS
    cv2.putText(frame, f"FPS: {fps:.1f}", (w - 150, 40),
               cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)


# ==============================================================================
# 🎬 MAIN
# ==============================================================================
def main():
    print("=" * 70)
    print("🤚 Simple Hand Gesture Detection (OpenCV Only)")
    print("=" * 70)
    
    init_excel_log()
    
    if USE_WEBCAM:
        print("📹 Using webcam")
        cap = VideoStream(0).start()
    else:
        print(f"📹 Pi Camera: {RTSP_URL}")
        cap = VideoStream(RTSP_URL).start()
    
    time.sleep(2.0)
    print("✅ Connected!\n")
    print("Gestures: 👊 FIST | ✌️ PEACE | ✋ OPEN HAND")
    print(f"📊 Logging: {EXCEL_FILE} (every {LOG_EVERY_N_SECONDS}s)")
    print("\n🖐️  INSTRUCTIONS:")
    print("  - Hold your hand in front of camera")
    print("  - Keep hand on RIGHT side of screen")
    print("  - Plain background works best")
    print("\nPress 'q' to quit\n")
    
    detector = SimpleGestureDetector()
    
    frame_count = 0
    fps_start = time.time()
    fps = 0
    last_log_time = time.time()
    
    current_gesture = "None"
    current_fingers = 0
    
    print("🎥 Detection started...\n")
    
    try:
        while True:
            frame = cap.read()
            if frame is None:
                continue
            
            # Mirror for easier interaction
            frame = cv2.flip(frame, 1)
            
            # ROI (Region of Interest) - Right side of frame
            h, w = frame.shape[:2]
            roi = frame[0:h, int(w/2):w]  # Right half
            
            # Detect skin
            skin_mask = detector.detect_skin(roi)
            
            # Find contours
            contours, _ = cv2.findContours(skin_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            
            finger_count = 0
            
            if contours:
                # Get largest contour (assume it's the hand)
                max_contour = max(contours, key=cv2.contourArea)
                
                # Only process if contour is large enough
                if cv2.contourArea(max_contour) > 5000:
                    # Draw contour on ROI
                    cv2.drawContours(roi, [max_contour], -1, (0, 255, 0), 2)
                    
                    # Get convex hull
                    hull = cv2.convexHull(max_contour, returnPoints=False)
                    
                    # Get convexity defects
                    if len(hull) > 3 and len(max_contour) > 3:
                        defects = cv2.convexityDefects(max_contour, hull)
                        finger_count = detector.count_fingers(max_contour, defects)
                        
                        # Draw hull
                        hull_points = cv2.convexHull(max_contour)
                        cv2.drawContours(roi, [hull_points], -1, (255, 0, 0), 2)
            
            # Recognize gesture
            current_gesture = detector.recognize_gesture(finger_count)
            current_fingers = finger_count
            
            # Draw ROI boundary
            cv2.rectangle(frame, (int(w/2), 0), (w, h), (255, 255, 0), 2)
            cv2.putText(frame, "Hand Area -->", (int(w/2) + 10, 30),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
            
            # Logging
            current_time = time.time()
            if (current_time - last_log_time) >= LOG_EVERY_N_SECONDS:
                if finger_count > 0:  # Only log when hand detected
                    log_to_excel(current_gesture, current_fingers)
                last_log_time = current_time
            
            # FPS
            frame_count += 1
            if frame_count % 30 == 0:
                fps = 30 / (time.time() - fps_start)
                fps_start = time.time()
            
            # Draw UI
            draw_info(frame, current_gesture, current_fingers, fps)
            
            # Display
            display_frame = cv2.resize(frame, (DISPLAY_WIDTH, DISPLAY_HEIGHT))
            cv2.imshow('🤚 Simple Gesture Detection', display_frame)
            
            # Quit
            if cv2.waitKey(1) & 0xFF == ord('q'):
                print("\n👋 Stopping...")
                if finger_count > 0:
                    log_to_excel(current_gesture, current_fingers)
                break
    
    except KeyboardInterrupt:
        print("\n⚠️  Interrupted")
    
    finally:
        print("🧹 Cleanup...")
        cap.stop()
        cv2.destroyAllWindows()
        print(f"✅ Done! Check {EXCEL_FILE}\n")


if __name__ == "__main__":
    main()
