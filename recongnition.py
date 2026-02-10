#!/usr/bin/env python3
import cv2
import socket
import time
import threading
import requests
import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from imutils.video import VideoStream
from options import Options

# Configuration
PI_IP = "100.109.95.64"
RTSP_URL = "rtsp://100.109.95.64:8554/feeder"
UDP_PORT = 5005
MIN_CONFIDENCE = 0.6
RESIZE_WIDTH = 640
LOG_FILE = "recognition_log.xlsx"  # 🆕 Excel log file

# Global state for threading
latest_frame = None
cached_results = []
is_running = True
new_frame_available = False

# Socket & Session setup
sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
session = requests.Session()
opts = Options()
current_led_state = None

# 🆕 Logging lock to prevent concurrent writes
log_lock = threading.Lock()

# ─────────────────────────────────────────
# 🆕 EXCEL LOGGING FUNCTIONS
# ─────────────────────────────────────────
def init_excel_log():
    """Create Excel file with headers if it doesn't exist."""
    if not os.path.exists(LOG_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recognition Log"
        
        # Headers
        headers = ["#", "Timestamp", "Name", "Confidence", "Status", "LED"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        
        wb.save(LOG_FILE)
        print(f"📊 Excel log created: {LOG_FILE}")

def log_to_excel(name, confidence, led_state):
    """Append a recognition event to Excel log."""
    with log_lock:
        try:
            wb = openpyxl.load_workbook(LOG_FILE)
            ws = wb.active
            
            # Row data
            row_num = ws.max_row  # Current row count (excluding header)
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            status = "Recognized" if name != "Unknown" else "Unknown"
            led = "ON" if led_state else "OFF"
            
            ws.append([row_num, timestamp, name, f"{confidence:.3f}", status, led])
            
            # Color rows based on status
            row = ws.max_row
            if name != "Unknown":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                font = Font(color="276221")
            else:
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Red
                font = Font(color="9C0006")
            
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(LOG_FILE)
            
        except Exception as e:
            print(f"❌ Log Error: {e}")

# ─────────────────────────────────────────
# EXISTING FUNCTIONS (unchanged)
# ─────────────────────────────────────────
def control_led(should_be_on):
    global current_led_state
    if current_led_state != should_be_on:
        try:
            sock.sendto(b'LED_ON' if should_be_on else b'LED_OFF', (PI_IP, UDP_PORT))
            current_led_state = should_be_on
        except Exception as e:
            print(f"LED Error: {e}")

def processing_thread():
    """ Background thread for API calls """
    global latest_frame, cached_results, is_running, new_frame_available
    
    # 🆕 Track last logged result to avoid duplicate entries
    last_logged_name = None
    last_log_time = 0
    LOG_COOLDOWN = 3.0  # 🆕 Log same person max once every 3 seconds
    
    while is_running:
        if latest_frame is not None and new_frame_available:
            new_frame_available = False
            frame_to_proc = latest_frame.copy()
            
            try:
                # 1. Resize and Encode
                h, w = frame_to_proc.shape[:2]
                small_frame = cv2.resize(frame_to_proc, (RESIZE_WIDTH, int(h * RESIZE_WIDTH / w)))
                _, encoded = cv2.imencode('.jpg', small_frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
                
                # 2. API Detection
                detect_response = session.post(
                    opts.endpoint("vision/face"),
                    files={"image": encoded.tobytes()},
                    timeout=1
                ).json()
                
                predictions = detect_response.get('predictions', [])
                results = []
                
                if predictions:
                    scale_x = w / RESIZE_WIDTH
                    scale_y = h / (h * RESIZE_WIDTH / w)
                    
                    for pred in predictions:
                        x_min, y_min = int(pred['x_min'] * scale_x), int(pred['y_min'] * scale_y)
                        x_max, y_max = int(pred['x_max'] * scale_x), int(pred['y_max'] * scale_y)
                        
                        face = frame_to_proc[y_min:y_max, x_min:x_max]
                        if face.size == 0: continue
                        
                        # 3. API Recognition
                        _, f_enc = cv2.imencode('.jpg', face, [cv2.IMWRITE_JPEG_QUALITY, 70])
                        rec_res = session.post(
                            opts.endpoint("vision/face/recognize"),
                            files={"image": f_enc.tobytes()},
                            data={"min_confidence": MIN_CONFIDENCE},
                            timeout=1
                        ).json()
                        
                        name = "Unknown"
                        conf = 0
                        if rec_res.get("predictions"):
                            user = rec_res["predictions"][0]
                            name = user.get("userid", "Unknown")
                            conf = user.get("confidence", 0)
                        
                        results.append({
                            'bbox': (x_min, y_min, x_max, y_max),
                            'name': name,
                            'confidence': conf
                        })
                
                # Update shared results and LED
                cached_results = results
                any_recognized = any(r['name'] != "Unknown" for r in results) if results else False
                control_led(any_recognized)
                
                # 🆕 Log to Excel (with cooldown to avoid spam)
                now = time.time()
                for r in results:
                    name = r['name']
                    conf = r['confidence']
                    
                    # Only log if: different person OR cooldown passed
                    if name != last_logged_name or (now - last_log_time) >= LOG_COOLDOWN:
                        threading.Thread(
                            target=log_to_excel,
                            args=(name, conf, any_recognized),
                            daemon=True
                        ).start()
                        last_logged_name = name
                        last_log_time = now
                        print(f"📝 Logged: {name} ({conf:.3f})")
                
            except Exception as e:
                print(f"Proc Error: {e}")
        else:
            time.sleep(0.01)

def main():
    global latest_frame, is_running, new_frame_available
    
    # 🆕 Initialize Excel log
    init_excel_log()
    
    cap = VideoStream(RTSP_URL).start()
    time.sleep(2.0)
    
    t = threading.Thread(target=processing_thread, daemon=True)
    t.start()
    
    print("System Live. Press 'q' to quit.")
    print(f"📊 Logging to: {LOG_FILE}\n")

    while True:
        frame = cap.read()
        if frame is None: continue
        
        latest_frame = frame
        new_frame_available = True
        
        # Draw results (unchanged)
        for result in cached_results:
            x_min, y_min, x_max, y_max = result['bbox']
            name, conf = result['name'], result['confidence']
            color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
            
            cv2.rectangle(frame, (x_min, y_min), (x_max, y_max), color, 2)
            label = f"{name} {conf:.2f}" if name != "Unknown" else "Unknown"
            cv2.putText(frame, label, (x_min, y_min - 10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
        
        cv2.imshow('Face Recognition (High Speed)', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    is_running = False
    t.join()
    control_led(False)
    cap.stop()
    cv2.destroyAllWindows()
    print(f"\n✅ Done! Log saved to: {LOG_FILE}")

if __name__ == "__main__":
    main()